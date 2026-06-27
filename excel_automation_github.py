from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

# create the workbook
workbook = Workbook()

# create the sheet in the excel
sheet = workbook.active

# add data into columns & rows
sheet["A1"] = "Name"
sheet["B1"] = "Marks"

sheet["A2"] = "Rahul"
sheet["B2"] = 20

sheet["A3"] = "Pavan"
sheet["B3"] = 99

sheet["A4"] = "Raj"
sheet["B4"] = 100

# save the excel file
workbook.save("std_marks.xlsx")

# load existing excel into memory
workbook = load_workbook("std_marks.xlsx")

# active current sheet
marks_sheet = workbook.active

total_std_marks = []

# get the marks from excel
for row in range(2, marks_sheet.max_row+1):
    marks = marks_sheet[f"B{row}"].value
    total_std_marks.append(marks)

# calculate total, avg, highest, lowest
total = sum(total_std_marks)
average = total / len(total_std_marks)
highest = max(total_std_marks)
lowest = min(total_std_marks)

# create new sheet 
marks_report = workbook.create_sheet("marks_report")

# add data into new sheet
marks_report["A1"] = "Total marks"
marks_report["B1"] = total

marks_report["A2"] = "Average"
marks_report["B2"] = average

marks_report["A3"] = "Highest"
marks_report["B3"] = average

marks_report["A4"] = "Lowest"
marks_report["B4"] = lowest

# data bold of excel
for row in range(1, marks_sheet.max_row+1):
    marks_report[f"A{row}"].font = Font(bold=True)

# add color to data of excel
yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
for row in range(1, marks_sheet.max_row+1):
    marks_report[f"A{row}"].fill = yellow_fill

# change width of excel columns
for col in range(1, 100):
    letter = get_column_letter(col)
    marks_sheet.column_dimensions[letter].width = 20

# change the number data format
marks_report["B1"].number_format = "0.000"
marks_report["B2"].number_format = "$0.00"
marks_report["B3"].number_format = "0.0%"

# create piechart object
chart = PieChart()

# get reference of data
data = Reference(
    marks_sheet,
    min_col=2,
    min_row=1,
    max_row=4
)

# add data to chart
chart.add_data(
    data,
    titles_from_data=True
)

# get reference of categories
categories = Reference(
    marks_sheet,
    min_col=1,
    min_row=2,
    max_row=4
)

# add categories in chart
chart.set_categories(categories)

# add chart in sheet
marks_sheet.add_chart(
    chart,
    "D2"
)

# save changes in the new excel
workbook.save("std_marks_summary.xlsx")
